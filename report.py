#!/usr/bin/env python

import utils
import sys
import getpass
import argparse

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Fill, NamedStyle
from toil.job import Job
from ddb import configuration
from ddb_ngsflow import pipeline
from variantstore import Variant
from variantstore import SampleVariant
from coveragestore import SampleCoverage
from cassandra.cqlengine import connection
from cassandra.auth import PlainTextAuthProvider


def process_sample(job, config, sample, samples, addresses, authenticator, thresholds, callers):
    job.fileStore.logToMaster("Retrieving data for sample {}\n".format(sample))
    job.fileStore.logToMaster("Retrieving coverage data from database\n")
    connection.setup(addresses, "coveragestore", auth_provider=authenticator)

    report_data = dict()
    filtered_variant_data = defaultdict(list)
    off_target_amplicon_counts = defaultdict(int)
    target_amplicon_coverage = dict()

    iterated = 0
    passing_variants = 0
    filtered_low_freq = 0
    filtered_low_depth = 0
    filtered_off_target = 0

    for library in samples[sample]:
        report_panel_path = "/mnt/shared-data/ddb-configs/disease_panels/{}/{}" \
                            "".format(samples[sample][library]['panel'], samples[sample][library]['report'])

        job.fileStore.logToMaster("{}: processing amplicons from file {}".format(library, report_panel_path))
        target_amplicons = utils.get_target_amplicons(report_panel_path)

        for amplicon in target_amplicons:
            coverage_data = SampleCoverage.objects.timeout(None).filter(
                SampleCoverage.sample == samples[sample][library]['sample_name'],
                SampleCoverage.amplicon == amplicon,
                SampleCoverage.run_id == samples[sample][library]['run_id'],
                SampleCoverage.library_name == samples[sample][library]['library_name'],
                SampleCoverage.program_name == "sambamba"
            )
            ordered_amplicons = coverage_data.order_by('amplicon', 'run_id').limit(coverage_data.count() + 1000)
            for result in ordered_amplicons:
                target_amplicon_coverage[amplicon] = result

        job.fileStore.logToMaster("{}: retrieving variants".format(library))
        variants = SampleVariant.objects.timeout(None).filter(
            SampleVariant.reference_genome == config['genome_version'],
            SampleVariant.sample == samples[sample][library]['sample_name'],
            SampleVariant.run_id == samples[sample][library]['run_id'],
            SampleVariant.library_name == samples[sample][library]['library_name'],
            SampleVariant.max_maf_all <= thresholds['max_maf']
        ).allow_filtering()

        num_var = variants.count()
        ordered = variants.order_by('library_name', 'chr', 'pos', 'ref', 'alt').limit(variants.count() + 1000)
        job.fileStore.logToMaster("{}: retrieved {} variants from database\n".format(library, num_var))
        job.fileStore.logToMaster("{}: classifying and filtering variants\n".format(library))

        for variant in ordered:
            iterated += 1
            if variant.amplicon_data['amplicon'] is 'None':
                filtered_off_target += 1
                off_target_amplicon_counts[variant.amplicon_data['amplicon']] += 1
            else:
                amplicons = variant.amplicon_data['amplicon'].split(',')
                assignable = 0
                for amplicon in amplicons:
                    if amplicon in target_amplicons:
                        assignable += 1
                        break
                if assignable:
                    # match_variants = Variant.objects.timeout(None).filter(
                    #     Variant.reference_genome == config['genome_version'],
                    #     Variant.chr == variant.chr,
                    #     Variant.pos == variant.pos,
                    #     Variant.ref == variant.ref,
                    #     Variant.alt == variant.alt
                    # ).allow_filtering()
                    #
                    # num_matches = match_variants.count()
                    # ordered_var = match_variants.order_by('sample', 'library_name', 'run_id').limit(num_matches + 1000)
                    # vafs = list()
                    # for var in ordered_var:
                    #     vaf = var.max_som_aaf
                    #     vafs.append(vaf)
                    # variant.vaf_median = np.median(vafs)
                    # variant.vaf_std_dev = np.std(vafs)

                    # Putting in to Tier1 based on COSMIC
                    if variant.cosmic_ids:
                        if variant.max_som_aaf < thresholds['min_saf']:
                            filtered_variant_data['tier1_fail_variants'].append(variant)
                            filtered_low_freq += 1
                        elif variant.max_depth < thresholds['depth']:
                            filtered_variant_data['tier1_fail_variants'].append(variant)
                            filtered_low_depth += 1
                        else:
                            filtered_variant_data['tier1_pass_variants'].append(variant)
                            passing_variants += 1
                        continue

                    # Putting in to Tier1 based on ClinVar not being None or Benign
                    if variant.clinvar_data['pathogenic'] != 'None':
                        if variant.clinvar_data['pathogenic'] != 'benign':
                            if variant.clinvar_data['pathogenic'] != 'likely-benign':
                                if variant.max_som_aaf < thresholds['min_saf']:
                                    filtered_variant_data['tier1_fail_variants'].append(variant)
                                    filtered_low_freq += 1
                                elif variant.max_depth < thresholds['depth']:
                                    filtered_variant_data['tier1_fail_variants'].append(variant)
                                    filtered_low_depth += 1
                                else:
                                    filtered_variant_data['tier1_pass_variants'].append(variant)
                                    passing_variants += 1
                                continue

                    if variant.severity == 'MED' or variant.severity == 'HIGH':
                        if variant.max_som_aaf < thresholds['min_saf']:
                            filtered_variant_data['tier3_fail_variants'].append(variant)
                            filtered_low_freq += 1
                        elif variant.max_depth < thresholds['depth']:
                            filtered_variant_data['tier3_fail_variants'].append(variant)
                            filtered_low_depth += 1
                        else:
                            filtered_variant_data['tier3_pass_variants'].append(variant)
                            passing_variants += 1
                        continue
                    else:
                        if variant.max_som_aaf < thresholds['min_saf']:
                            filtered_variant_data['tier4_fail_variants'].append(variant)
                            filtered_low_freq += 1
                        elif variant.max_depth < thresholds['depth']:
                            filtered_variant_data['tier4_fail_variants'].append(variant)
                            filtered_low_depth += 1
                        else:
                            filtered_variant_data['tier4_pass_variants'].append(variant)
                            passing_variants += 1
                        continue
                else:
                    filtered_off_target += 1
                    off_target_amplicon_counts[variant.amplicon_data['amplicon']] += 1

        job.fileStore.logToMaster("{}: iterated through {} variants\n".format(library, iterated))
        job.fileStore.logToMaster("{}: filtered {} off-target variants\n".format(library, filtered_off_target))
        job.fileStore.logToMaster("{}: filtered {} low-freq variants\n".format(library, filtered_low_freq))
        job.fileStore.logToMaster("{}: filtered {} low-depth variants\n".format(library, filtered_low_depth))

        job.fileStore.logToMaster("{}: passing {} tier 1 and 2 variants"
                                  "\n".format(library, len(filtered_variant_data['tier1_pass_variants'])))
        job.fileStore.logToMaster("{}: passing {} tier3 variants"
                                  "\n".format(library, len(filtered_variant_data['vus_pass_variants'])))
        job.fileStore.logToMaster("{}: passing {} tier 4 variants"
                                  "\n".format(library, len(filtered_variant_data['tier4_pass_variants'])))

    report_data['variants'] = filtered_variant_data
    report_data['coverage'] = target_amplicon_coverage

    report_name = "{}.xlsx".format(sample)
    wb = Workbook()
    coverage_sheet = wb.create_sheet(title="Coverage")
    tier1_sheet = wb.create_sheet(title="Tier1 and 2 Pass")
    tier3_sheet = wb.create_sheet(title="Tier3 Pass")
    tier4_sheet = wb.create_sheet(title="Tier4 Pass")
    tier1_fail_sheet = wb.create_sheet(title="Tier1 and 2 Fail")
    tier3_fail_sheet = wb.create_sheet(title="Tier3 Fail")
    tier4_fail_sheet = wb.create_sheet(title="Tier4 Fail")

    libraries = list()
    report_templates = list()
    run_id = ""
    for library in samples[sample]:
        libraries.append(samples[sample][library]['library_name'])
        report_templates.append(samples[sample][library]['report'])
        run_id = samples[sample][library]['run_id']
    lib_string = " | ".join(libraries)
    reports_string = " | ".join(report_templates)

    coverage_sheet.cell(row=1, column=1, value="Sample")
    coverage_sheet.cell(row=1, column=2, value="{}".format(sample))

    coverage_sheet.cell(row=2, column=1, value="Libraries")
    coverage_sheet.cell(row=2, column=2, value="{}".format(lib_string))

    coverage_sheet.cell(row=3, column=1, value="Run ID")
    coverage_sheet.cell(row=3, column=2, value="{}".format(run_id))

    coverage_sheet.cell(row=4, column=1, value="Reporting Templates")
    coverage_sheet.cell(row=4, column=2, value="{}".format(reports_string))

    coverage_sheet.cell(row=5, column=1, value="Minimum Reportable Somatic Allele Frequency")
    coverage_sheet.cell(row=5, column=2, value="{}".format(thresholds['min_saf']))

    coverage_sheet.cell(row=6, column=1, value="Minimum Amplicon Depth")
    coverage_sheet.cell(row=6, column=2, value="{}".format(thresholds['depth']))

    coverage_sheet.cell(row=7, column=1, value="Maximum Population Allele Frequency")
    coverage_sheet.cell(row=7, column=2, value="{}".format(thresholds['max_maf']))

    coverage_sheet.cell(row=8, column=1, value="Sample")
    coverage_sheet.cell(row=8, column=2, value="Library")
    coverage_sheet.cell(row=8, column=3, value="Amplicon")
    coverage_sheet.cell(row=8, column=4, value="Num Reads")
    coverage_sheet.cell(row=8, column=5, value="Coverage")

    error = NamedStyle(name="error")
    warning = NamedStyle(name="warning")
    good = NamedStyle(name="good")

    error.fill = PatternFill("solid", fgColour="FFC7CE")
    warning.fill = PatternFill("solid", fgColour="FDC478")
    good.fill = PatternFill("solid", fgColour="3ED626")

    wb.add_named_style(error)
    wb.add_named_style(warning)
    wb.add_named_style(good)

    row_num = 9
    for amplicon in report_data['coverage']:
        coverage_sheet.cell(row=row_num, column=1, value="{}".format(report_data['coverage'][amplicon].sample))
        coverage_sheet.cell(row=row_num, column=2, value="{}".format(report_data['coverage'][amplicon].library_name))
        coverage_sheet.cell(row=row_num, column=3, value="{}".format(report_data['coverage'][amplicon].amplicon))
        coverage_sheet.cell(row=row_num, column=4, value="{}".format(report_data['coverage'][amplicon].num_reads))
        coverage_sheet.cell(row=row_num, column=5, value="{}".format(report_data['coverage'][amplicon].mean_coverage))

        num_cell = coverage_sheet.cell(row=row_num, column=4)
        cov_cell = coverage_sheet.cell(row=row_num, column=5)

        if report_data['coverage'][amplicon].mean_coverage < 250:
            num_cell.style = error
            cov_cell.style = error
        elif report_data['coverage'][amplicon].mean_coverage < 500:
            num_cell.style = warning
            cov_cell.style = warning
        else:
            num_cell.style = good
            cov_cell.style = good

        row_num += 1

    ####################################################################################################################

    tier1_sheet.cell(row=1, column=1, value="Sample")
    tier1_sheet.cell(row=1, column=2, value="Library")
    tier1_sheet.cell(row=1, column=3, value="Gene")
    tier1_sheet.cell(row=1, column=4, value="Amplicon")
    tier1_sheet.cell(row=1, column=5, value="Ref")
    tier1_sheet.cell(row=1, column=6, value="Alt")
    tier1_sheet.cell(row=1, column=7, value="Codon")
    tier1_sheet.cell(row=1, column=8, value="AA")
    tier1_sheet.cell(row=1, column=9, value="Max Caller Somatic VAF")
    tier1_sheet.cell(row=1, column=10, value="Callers")
    tier1_sheet.cell(row=1, column=11, value="COSMIC IDs")
    tier1_sheet.cell(row=1, column=12, value="Num COSMIC Samples")
    tier1_sheet.cell(row=1, column=13, value="COSMIC AA")
    tier1_sheet.cell(row=1, column=14, value="Clinvar Significance")
    tier1_sheet.cell(row=1, column=15, value="Clinvar HGVS")
    tier1_sheet.cell(row=1, column=16, value="Clinvar Disease")
    tier1_sheet.cell(row=1, column=17, value="Coverage")
    tier1_sheet.cell(row=1, column=18, value="Num Reads")
    tier1_sheet.cell(row=1, column=19, value="Impact")
    tier1_sheet.cell(row=1, column=20, value="Severity")
    tier1_sheet.cell(row=1, column=21, value="Maximum Population AF")
    tier1_sheet.cell(row=1, column=22, value="Min Caller Depth")
    tier1_sheet.cell(row=1, column=23, value="Max Caller Depth")
    tier1_sheet.cell(row=1, column=24, value="Chrom")
    tier1_sheet.cell(row=1, column=25, value="Start")
    tier1_sheet.cell(row=1, column=26, value="End")
    tier1_sheet.cell(row=1, column=27, value="rsIDs")

    col = 28
    if 'mutect' in callers:
        tier1_sheet.cell(row=1, column=col, value="MuTect_AF")
        col += 1

    if 'vardict' in callers:
        tier1_sheet.cell(row=1, column=col, value="VarDict_AF")
        col += 1

    if 'freebayes' in callers:
        tier1_sheet.cell(row=1, column=col, value="FreeBayes_AF")
        col += 1

    if 'scalpel' in callers:
        tier1_sheet.cell(row=1, column=col, value="Scalpel_AF")
        col += 1

    if 'platypus' in callers:
        tier1_sheet.cell(row=1, column=col, value="Platypus_AF")
        col += 1

    if 'pindel' in callers:
        tier1_sheet.cell(row=1, column=col, value="Pindel_AF")
        col += 1

    row = 2
    for variant in report_data['variants']['tier1_pass_variants']:
        tier1_sheet.cell(row=row, column=1, value="{}".format(variant.sample))
        tier1_sheet.cell(row=row, column=2, value="{}".format(variant.library_name))
        tier1_sheet.cell(row=row, column=3, value="{}".format(variant.gene))
        tier1_sheet.cell(row=row, column=4, value="{}".format(variant.amplicon_data['amplicon']))
        tier1_sheet.cell(row=row, column=5, value="{}".format(variant.ref))
        tier1_sheet.cell(row=row, column=6, value="{}".format(variant.alt))
        tier1_sheet.cell(row=row, column=7, value="{}".format(variant.codon_change))
        tier1_sheet.cell(row=row, column=8, value="{}".format(variant.aa_change))
        tier1_sheet.cell(row=row, column=9, value="{}".format(variant.max_som_aaf))
        tier1_sheet.cell(row=row, column=10, value="{}".format(",".join(variant.callers) or None))
        tier1_sheet.cell(row=row, column=11, value="{}".format(",".join(variant.cosmic_ids) or None))
        tier1_sheet.cell(row=row, column=12, value="{}".format(variant.cosmic_data['num_samples']))
        tier1_sheet.cell(row=row, column=13, value="{}".format(variant.cosmic_data['aa']))
        tier1_sheet.cell(row=row, column=14, value="{}".format(variant.clinvar_data['significance']))
        tier1_sheet.cell(row=row, column=15, value="{}".format(variant.clinvar_data['hgvs']))
        tier1_sheet.cell(row=row, column=16, value="{}".format(variant.clinvar_data['disease']))
        tier1_sheet.cell(row=row, column=17,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['mean_coverage']))
        tier1_sheet.cell(row=row, column=18,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['num_reads']))
        tier1_sheet.cell(row=row, column=19, value="{}".format(variant.impact))
        tier1_sheet.cell(row=row, column=20, value="{}".format(variant.severity))
        tier1_sheet.cell(row=row, column=21, value="{}".format(variant.max_maf_all))
        tier1_sheet.cell(row=row, column=22, value="{}".format(variant.min_depth))
        tier1_sheet.cell(row=row, column=23, value="{}".format(variant.max_depth))
        tier1_sheet.cell(row=row, column=24, value="{}".format(variant.chr))
        tier1_sheet.cell(row=row, column=25, value="{}".format(variant.pos))
        tier1_sheet.cell(row=row, column=26, value="{}".format(variant.end))
        tier1_sheet.cell(row=row, column=27, value="{}".format(",".join(variant.rs_ids)))

        col = 28
        if 'mutect' in callers:
            tier1_sheet.cell(row=row, column=col, value="{}".format(variant.mutect.get('AAF') or None))
            col += 1

        if 'vardict' in callers:
            tier1_sheet.cell(row=row, column=col, value="{}".format(variant.vardict.get('AAF') or None))
            col += 1

        if 'freebayes' in callers:
            tier1_sheet.cell(row=row, column=col, value="{}".format(variant.freebayes.get('AAF') or None))
            col += 1

        if 'scalpel' in callers:
            tier1_sheet.cell(row=row, column=col, value="{}".format(variant.scalpel.get('AAF') or None))
            col += 1

        if 'platypus' in callers:
            tier1_sheet.cell(row=row, column=col, value="{}".format(variant.platypus.get('AAF') or None))
            col += 1

        if 'pindel' in callers:
            tier1_sheet.cell(row=row, column=col, value="{}".format(variant.pindel.get('AAF') or None))
            col += 1

        row += 1

    ####################################################################################################################

    tier3_sheet.cell(row=1, column=1, value="Sample")
    tier3_sheet.cell(row=1, column=2, value="Library")
    tier3_sheet.cell(row=1, column=3, value="Gene")
    tier3_sheet.cell(row=1, column=4, value="Amplicon")
    tier3_sheet.cell(row=1, column=5, value="Ref")
    tier3_sheet.cell(row=1, column=6, value="Alt")
    tier3_sheet.cell(row=1, column=7, value="Codon")
    tier3_sheet.cell(row=1, column=8, value="AA")
    tier3_sheet.cell(row=1, column=9, value="Max Caller Somatic VAF")
    tier3_sheet.cell(row=1, column=10, value="Callers")
    tier3_sheet.cell(row=1, column=11, value="COSMIC IDs")
    tier3_sheet.cell(row=1, column=12, value="Num COSMIC Samples")
    tier3_sheet.cell(row=1, column=13, value="COSMIC AA")
    tier3_sheet.cell(row=1, column=14, value="Clinvar Significance")
    tier3_sheet.cell(row=1, column=15, value="Clinvar HGVS")
    tier3_sheet.cell(row=1, column=16, value="Clinvar Disease")
    tier3_sheet.cell(row=1, column=17, value="Coverage")
    tier3_sheet.cell(row=1, column=18, value="Num Reads")
    tier3_sheet.cell(row=1, column=19, value="Impact")
    tier3_sheet.cell(row=1, column=20, value="Severity")
    tier3_sheet.cell(row=1, column=21, value="Maximum Population AF")
    tier3_sheet.cell(row=1, column=22, value="Min Caller Depth")
    tier3_sheet.cell(row=1, column=23, value="Max Caller Depth")
    tier3_sheet.cell(row=1, column=24, value="Chrom")
    tier3_sheet.cell(row=1, column=25, value="Start")
    tier3_sheet.cell(row=1, column=26, value="End")
    tier3_sheet.cell(row=1, column=27, value="rsIDs")

    col = 28
    if 'mutect' in callers:
        tier3_sheet.cell(row=1, column=col, value="MuTect_AF")
        col += 1

    if 'vardict' in callers:
        tier3_sheet.cell(row=1, column=col, value="VarDict_AF")
        col += 1

    if 'freebayes' in callers:
        tier3_sheet.cell(row=1, column=col, value="FreeBayes_AF")
        col += 1

    if 'scalpel' in callers:
        tier3_sheet.cell(row=1, column=col, value="Scalpel_AF")
        col += 1

    if 'platypus' in callers:
        tier3_sheet.cell(row=1, column=col, value="Platypus_AF")
        col += 1

    if 'pindel' in callers:
        tier3_sheet.cell(row=1, column=col, value="Pindel_AF")
        col += 1

    row = 2
    for variant in report_data['variants']['tier3_pass_variants']:
        tier3_sheet.cell(row=row, column=1, value="{}".format(variant.sample))
        tier3_sheet.cell(row=row, column=2, value="{}".format(variant.library_name))
        tier3_sheet.cell(row=row, column=3, value="{}".format(variant.gene))
        tier3_sheet.cell(row=row, column=4, value="{}".format(variant.amplicon_data['amplicon']))
        tier3_sheet.cell(row=row, column=5, value="{}".format(variant.ref))
        tier3_sheet.cell(row=row, column=6, value="{}".format(variant.alt))
        tier3_sheet.cell(row=row, column=7, value="{}".format(variant.codon_change))
        tier3_sheet.cell(row=row, column=8, value="{}".format(variant.aa_change))
        tier3_sheet.cell(row=row, column=9, value="{}".format(variant.max_som_aaf))
        tier3_sheet.cell(row=row, column=10, value="{}".format(",".join(variant.callers) or None))
        tier3_sheet.cell(row=row, column=11, value="{}".format(",".join(variant.cosmic_ids) or None))
        tier3_sheet.cell(row=row, column=12, value="{}".format(variant.cosmic_data['num_samples']))
        tier3_sheet.cell(row=row, column=13, value="{}".format(variant.cosmic_data['aa']))
        tier3_sheet.cell(row=row, column=14, value="{}".format(variant.clinvar_data['significance']))
        tier3_sheet.cell(row=row, column=15, value="{}".format(variant.clinvar_data['hgvs']))
        tier3_sheet.cell(row=row, column=16, value="{}".format(variant.clinvar_data['disease']))
        tier3_sheet.cell(row=row, column=17,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['mean_coverage']))
        tier3_sheet.cell(row=row, column=18,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['num_reads']))
        tier3_sheet.cell(row=row, column=19, value="{}".format(variant.impact))
        tier3_sheet.cell(row=row, column=20, value="{}".format(variant.severity))
        tier3_sheet.cell(row=row, column=21, value="{}".format(variant.max_maf_all))
        tier3_sheet.cell(row=row, column=22, value="{}".format(variant.min_depth))
        tier3_sheet.cell(row=row, column=23, value="{}".format(variant.max_depth))
        tier3_sheet.cell(row=row, column=24, value="{}".format(variant.chr))
        tier3_sheet.cell(row=row, column=25, value="{}".format(variant.pos))
        tier3_sheet.cell(row=row, column=26, value="{}".format(variant.end))
        tier3_sheet.cell(row=row, column=27, value="{}".format(",".join(variant.rs_ids)))

        col = 28
        if 'mutect' in callers:
            tier3_sheet.cell(row=row, column=col, value="{}".format(variant.mutect.get('AAF') or None))
            col += 1

        if 'vardict' in callers:
            tier3_sheet.cell(row=row, column=col, value="{}".format(variant.vardict.get('AAF') or None))
            col += 1

        if 'freebayes' in callers:
            tier3_sheet.cell(row=row, column=col, value="{}".format(variant.freebayes.get('AAF') or None))
            col += 1

        if 'scalpel' in callers:
            tier3_sheet.cell(row=row, column=col, value="{}".format(variant.scalpel.get('AAF') or None))
            col += 1

        if 'platypus' in callers:
            tier3_sheet.cell(row=row, column=col, value="{}".format(variant.platypus.get('AAF') or None))
            col += 1

        if 'pindel' in callers:
            tier3_sheet.cell(row=row, column=col, value="{}".format(variant.pindel.get('AAF') or None))
            col += 1

    ####################################################################################################################

    tier4_sheet.cell(row=1, column=1, value="Sample")
    tier4_sheet.cell(row=1, column=2, value="Library")
    tier4_sheet.cell(row=1, column=3, value="Gene")
    tier4_sheet.cell(row=1, column=4, value="Amplicon")
    tier4_sheet.cell(row=1, column=5, value="Ref")
    tier4_sheet.cell(row=1, column=6, value="Alt")
    tier4_sheet.cell(row=1, column=7, value="Codon")
    tier4_sheet.cell(row=1, column=8, value="AA")
    tier4_sheet.cell(row=1, column=9, value="Max Caller Somatic VAF")
    tier4_sheet.cell(row=1, column=10, value="Callers")
    tier4_sheet.cell(row=1, column=11, value="COSMIC IDs")
    tier4_sheet.cell(row=1, column=12, value="Num COSMIC Samples")
    tier4_sheet.cell(row=1, column=13, value="COSMIC AA")
    tier4_sheet.cell(row=1, column=14, value="Clinvar Significance")
    tier4_sheet.cell(row=1, column=15, value="Clinvar HGVS")
    tier4_sheet.cell(row=1, column=16, value="Clinvar Disease")
    tier4_sheet.cell(row=1, column=17, value="Coverage")
    tier4_sheet.cell(row=1, column=18, value="Num Reads")
    tier4_sheet.cell(row=1, column=19, value="Impact")
    tier4_sheet.cell(row=1, column=20, value="Severity")
    tier4_sheet.cell(row=1, column=21, value="Maximum Population AF")
    tier4_sheet.cell(row=1, column=22, value="Min Caller Depth")
    tier4_sheet.cell(row=1, column=23, value="Max Caller Depth")
    tier4_sheet.cell(row=1, column=24, value="Chrom")
    tier4_sheet.cell(row=1, column=25, value="Start")
    tier4_sheet.cell(row=1, column=26, value="End")
    tier4_sheet.cell(row=1, column=27, value="rsIDs")

    col = 28
    if 'mutect' in callers:
        tier4_sheet.cell(row=1, column=col, value="MuTect_AF")
        col += 1

    if 'vardict' in callers:
        tier4_sheet.cell(row=1, column=col, value="VarDict_AF")
        col += 1

    if 'freebayes' in callers:
        tier4_sheet.cell(row=1, column=col, value="FreeBayes_AF")
        col += 1

    if 'scalpel' in callers:
        tier4_sheet.cell(row=1, column=col, value="Scalpel_AF")
        col += 1

    if 'platypus' in callers:
        tier4_sheet.cell(row=1, column=col, value="Platypus_AF")
        col += 1

    if 'pindel' in callers:
        tier4_sheet.cell(row=1, column=col, value="Pindel_AF")
        col += 1

    row = 2
    for variant in report_data['variants']['tier4_pass_variants']:
        tier4_sheet.cell(row=row, column=1, value="{}".format(variant.sample))
        tier4_sheet.cell(row=row, column=2, value="{}".format(variant.library_name))
        tier4_sheet.cell(row=row, column=3, value="{}".format(variant.gene))
        tier4_sheet.cell(row=row, column=4, value="{}".format(variant.amplicon_data['amplicon']))
        tier4_sheet.cell(row=row, column=5, value="{}".format(variant.ref))
        tier4_sheet.cell(row=row, column=6, value="{}".format(variant.alt))
        tier4_sheet.cell(row=row, column=7, value="{}".format(variant.codon_change))
        tier4_sheet.cell(row=row, column=8, value="{}".format(variant.aa_change))
        tier4_sheet.cell(row=row, column=9, value="{}".format(variant.max_som_aaf))
        tier4_sheet.cell(row=row, column=10, value="{}".format(",".join(variant.callers) or None))
        tier4_sheet.cell(row=row, column=11, value="{}".format(",".join(variant.cosmic_ids) or None))
        tier4_sheet.cell(row=row, column=12, value="{}".format(variant.cosmic_data['num_samples']))
        tier4_sheet.cell(row=row, column=13, value="{}".format(variant.cosmic_data['aa']))
        tier4_sheet.cell(row=row, column=14, value="{}".format(variant.clinvar_data['significance']))
        tier4_sheet.cell(row=row, column=15, value="{}".format(variant.clinvar_data['hgvs']))
        tier4_sheet.cell(row=row, column=16, value="{}".format(variant.clinvar_data['disease']))
        tier4_sheet.cell(row=row, column=17,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['mean_coverage']))
        tier4_sheet.cell(row=row, column=18,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['num_reads']))
        tier4_sheet.cell(row=row, column=19, value="{}".format(variant.impact))
        tier4_sheet.cell(row=row, column=20, value="{}".format(variant.severity))
        tier4_sheet.cell(row=row, column=21, value="{}".format(variant.max_maf_all))
        tier4_sheet.cell(row=row, column=22, value="{}".format(variant.min_depth))
        tier4_sheet.cell(row=row, column=23, value="{}".format(variant.max_depth))
        tier4_sheet.cell(row=row, column=24, value="{}".format(variant.chr))
        tier4_sheet.cell(row=row, column=25, value="{}".format(variant.pos))
        tier4_sheet.cell(row=row, column=26, value="{}".format(variant.end))
        tier4_sheet.cell(row=row, column=27, value="{}".format(",".join(variant.rs_ids)))

        col = 28
        if 'mutect' in callers:
            tier4_sheet.cell(row=row, column=col, value="{}".format(variant.mutect.get('AAF') or None))
            col += 1

        if 'vardict' in callers:
            tier4_sheet.cell(row=row, column=col, value="{}".format(variant.vardict.get('AAF') or None))
            col += 1

        if 'freebayes' in callers:
            tier4_sheet.cell(row=row, column=col, value="{}".format(variant.freebayes.get('AAF') or None))
            col += 1

        if 'scalpel' in callers:
            tier4_sheet.cell(row=row, column=col, value="{}".format(variant.scalpel.get('AAF') or None))
            col += 1

        if 'platypus' in callers:
            tier4_sheet.cell(row=row, column=col, value="{}".format(variant.platypus.get('AAF') or None))
            col += 1

        if 'pindel' in callers:
            tier4_sheet.cell(row=row, column=col, value="{}".format(variant.pindel.get('AAF') or None))
            col += 1

    ####################################################################################################################

    tier1_fail_sheet.cell(row=1, column=1, value="Sample")
    tier1_fail_sheet.cell(row=1, column=2, value="Library")
    tier1_fail_sheet.cell(row=1, column=3, value="Gene")
    tier1_fail_sheet.cell(row=1, column=4, value="Amplicon")
    tier1_fail_sheet.cell(row=1, column=5, value="Ref")
    tier1_fail_sheet.cell(row=1, column=6, value="Alt")
    tier1_fail_sheet.cell(row=1, column=7, value="Codon")
    tier1_fail_sheet.cell(row=1, column=8, value="AA")
    tier1_fail_sheet.cell(row=1, column=9, value="Max Caller Somatic VAF")
    tier1_fail_sheet.cell(row=1, column=10, value="Callers")
    tier1_fail_sheet.cell(row=1, column=11, value="COSMIC IDs")
    tier1_fail_sheet.cell(row=1, column=12, value="Num COSMIC Samples")
    tier1_fail_sheet.cell(row=1, column=13, value="COSMIC AA")
    tier1_fail_sheet.cell(row=1, column=14, value="Clinvar Significance")
    tier1_fail_sheet.cell(row=1, column=15, value="Clinvar HGVS")
    tier1_fail_sheet.cell(row=1, column=16, value="Clinvar Disease")
    tier1_fail_sheet.cell(row=1, column=17, value="Coverage")
    tier1_fail_sheet.cell(row=1, column=18, value="Num Reads")
    tier1_fail_sheet.cell(row=1, column=19, value="Impact")
    tier1_fail_sheet.cell(row=1, column=20, value="Severity")
    tier1_fail_sheet.cell(row=1, column=21, value="Maximum Population AF")
    tier1_fail_sheet.cell(row=1, column=22, value="Min Caller Depth")
    tier1_fail_sheet.cell(row=1, column=23, value="Max Caller Depth")
    tier1_fail_sheet.cell(row=1, column=24, value="Chrom")
    tier1_fail_sheet.cell(row=1, column=25, value="Start")
    tier1_fail_sheet.cell(row=1, column=26, value="End")
    tier1_fail_sheet.cell(row=1, column=27, value="rsIDs")

    col = 28
    if 'mutect' in callers:
        tier1_fail_sheet.cell(row=1, column=col, value="MuTect_AF")
        col += 1

    if 'vardict' in callers:
        tier1_fail_sheet.cell(row=1, column=col, value="VarDict_AF")
        col += 1

    if 'freebayes' in callers:
        tier1_fail_sheet.cell(row=1, column=col, value="FreeBayes_AF")
        col += 1

    if 'scalpel' in callers:
        tier1_fail_sheet.cell(row=1, column=col, value="Scalpel_AF")
        col += 1

    if 'platypus' in callers:
        tier1_fail_sheet.cell(row=1, column=col, value="Platypus_AF")
        col += 1

    if 'pindel' in callers:
        tier1_fail_sheet.cell(row=1, column=col, value="Pindel_AF")
        col += 1

    row = 2
    for variant in report_data['variants']['tier1_fail_variants']:
        tier1_fail_sheet.cell(row=row, column=1, value="{}".format(variant.sample))
        tier1_fail_sheet.cell(row=row, column=2, value="{}".format(variant.library_name))
        tier1_fail_sheet.cell(row=row, column=3, value="{}".format(variant.gene))
        tier1_fail_sheet.cell(row=row, column=4, value="{}".format(variant.amplicon_data['amplicon']))
        tier1_fail_sheet.cell(row=row, column=5, value="{}".format(variant.ref))
        tier1_fail_sheet.cell(row=row, column=6, value="{}".format(variant.alt))
        tier1_fail_sheet.cell(row=row, column=7, value="{}".format(variant.codon_change))
        tier1_fail_sheet.cell(row=row, column=8, value="{}".format(variant.aa_change))
        tier1_fail_sheet.cell(row=row, column=9, value="{}".format(variant.max_som_aaf))
        tier1_fail_sheet.cell(row=row, column=10, value="{}".format(",".join(variant.callers) or None))
        tier1_fail_sheet.cell(row=row, column=11, value="{}".format(",".join(variant.cosmic_ids) or None))
        tier1_fail_sheet.cell(row=row, column=12, value="{}".format(variant.cosmic_data['num_samples']))
        tier1_fail_sheet.cell(row=row, column=13, value="{}".format(variant.cosmic_data['aa']))
        tier1_fail_sheet.cell(row=row, column=14, value="{}".format(variant.clinvar_data['significance']))
        tier1_fail_sheet.cell(row=row, column=15, value="{}".format(variant.clinvar_data['hgvs']))
        tier1_fail_sheet.cell(row=row, column=16, value="{}".format(variant.clinvar_data['disease']))
        tier1_fail_sheet.cell(row=row, column=17,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['mean_coverage']))
        tier1_fail_sheet.cell(row=row, column=18,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['num_reads']))
        tier1_fail_sheet.cell(row=row, column=19, value="{}".format(variant.impact))
        tier1_fail_sheet.cell(row=row, column=20, value="{}".format(variant.severity))
        tier1_fail_sheet.cell(row=row, column=21, value="{}".format(variant.max_maf_all))
        tier1_fail_sheet.cell(row=row, column=22, value="{}".format(variant.min_depth))
        tier1_fail_sheet.cell(row=row, column=23, value="{}".format(variant.max_depth))
        tier1_fail_sheet.cell(row=row, column=24, value="{}".format(variant.chr))
        tier1_fail_sheet.cell(row=row, column=25, value="{}".format(variant.pos))
        tier1_fail_sheet.cell(row=row, column=26, value="{}".format(variant.end))
        tier1_fail_sheet.cell(row=row, column=27, value="{}".format(",".join(variant.rs_ids)))

        col = 28
        if 'mutect' in callers:
            tier1_fail_sheet.cell(row=row, column=col, value="{}".format(variant.mutect.get('AAF') or None))
            col += 1

        if 'vardict' in callers:
            tier1_fail_sheet.cell(row=row, column=col, value="{}".format(variant.vardict.get('AAF') or None))
            col += 1

        if 'freebayes' in callers:
            tier1_fail_sheet.cell(row=row, column=col, value="{}".format(variant.freebayes.get('AAF') or None))
            col += 1

        if 'scalpel' in callers:
            tier1_fail_sheet.cell(row=row, column=col, value="{}".format(variant.scalpel.get('AAF') or None))
            col += 1

        if 'platypus' in callers:
            tier1_fail_sheet.cell(row=row, column=col, value="{}".format(variant.platypus.get('AAF') or None))
            col += 1

        if 'pindel' in callers:
            tier1_fail_sheet.cell(row=row, column=col, value="{}".format(variant.pindel.get('AAF') or None))
            col += 1

        row += 1

    ####################################################################################################################

    tier3_fail_sheet.cell(row=1, column=1, value="Sample")
    tier3_fail_sheet.cell(row=1, column=2, value="Library")
    tier3_fail_sheet.cell(row=1, column=3, value="Gene")
    tier3_fail_sheet.cell(row=1, column=4, value="Amplicon")
    tier3_fail_sheet.cell(row=1, column=5, value="Ref")
    tier3_fail_sheet.cell(row=1, column=6, value="Alt")
    tier3_fail_sheet.cell(row=1, column=7, value="Codon")
    tier3_fail_sheet.cell(row=1, column=8, value="AA")
    tier3_fail_sheet.cell(row=1, column=9, value="Max Caller Somatic VAF")
    tier3_fail_sheet.cell(row=1, column=10, value="Callers")
    tier3_fail_sheet.cell(row=1, column=11, value="COSMIC IDs")
    tier3_fail_sheet.cell(row=1, column=12, value="Num COSMIC Samples")
    tier3_fail_sheet.cell(row=1, column=13, value="COSMIC AA")
    tier3_fail_sheet.cell(row=1, column=14, value="Clinvar Significance")
    tier3_fail_sheet.cell(row=1, column=15, value="Clinvar HGVS")
    tier3_fail_sheet.cell(row=1, column=16, value="Clinvar Disease")
    tier3_fail_sheet.cell(row=1, column=17, value="Coverage")
    tier3_fail_sheet.cell(row=1, column=18, value="Num Reads")
    tier3_fail_sheet.cell(row=1, column=19, value="Impact")
    tier3_fail_sheet.cell(row=1, column=20, value="Severity")
    tier3_fail_sheet.cell(row=1, column=21, value="Maximum Population AF")
    tier3_fail_sheet.cell(row=1, column=22, value="Min Caller Depth")
    tier3_fail_sheet.cell(row=1, column=23, value="Max Caller Depth")
    tier3_fail_sheet.cell(row=1, column=24, value="Chrom")
    tier3_fail_sheet.cell(row=1, column=25, value="Start")
    tier3_fail_sheet.cell(row=1, column=26, value="End")
    tier3_fail_sheet.cell(row=1, column=27, value="rsIDs")

    col = 28
    if 'mutect' in callers:
        tier3_fail_sheet.cell(row=1, column=col, value="MuTect_AF")
        col += 1

    if 'vardict' in callers:
        tier3_fail_sheet.cell(row=1, column=col, value="VarDict_AF")
        col += 1

    if 'freebayes' in callers:
        tier3_fail_sheet.cell(row=1, column=col, value="FreeBayes_AF")
        col += 1

    if 'scalpel' in callers:
        tier3_fail_sheet.cell(row=1, column=col, value="Scalpel_AF")
        col += 1

    if 'platypus' in callers:
        tier3_fail_sheet.cell(row=1, column=col, value="Platypus_AF")
        col += 1

    if 'pindel' in callers:
        tier3_fail_sheet.cell(row=1, column=col, value="Pindel_AF")
        col += 1

    row = 2
    for variant in report_data['variants']['tier3_fail_variants']:
        tier3_fail_sheet.cell(row=row, column=1, value="{}".format(variant.sample))
        tier3_fail_sheet.cell(row=row, column=2, value="{}".format(variant.library_name))
        tier3_fail_sheet.cell(row=row, column=3, value="{}".format(variant.gene))
        tier3_fail_sheet.cell(row=row, column=4, value="{}".format(variant.amplicon_data['amplicon']))
        tier3_fail_sheet.cell(row=row, column=5, value="{}".format(variant.ref))
        tier3_fail_sheet.cell(row=row, column=6, value="{}".format(variant.alt))
        tier3_fail_sheet.cell(row=row, column=7, value="{}".format(variant.codon_change))
        tier3_fail_sheet.cell(row=row, column=8, value="{}".format(variant.aa_change))
        tier3_fail_sheet.cell(row=row, column=9, value="{}".format(variant.max_som_aaf))
        tier3_fail_sheet.cell(row=row, column=10, value="{}".format(",".join(variant.callers) or None))
        tier3_fail_sheet.cell(row=row, column=11, value="{}".format(",".join(variant.cosmic_ids) or None))
        tier3_fail_sheet.cell(row=row, column=12, value="{}".format(variant.cosmic_data['num_samples']))
        tier3_fail_sheet.cell(row=row, column=13, value="{}".format(variant.cosmic_data['aa']))
        tier3_fail_sheet.cell(row=row, column=14, value="{}".format(variant.clinvar_data['significance']))
        tier3_fail_sheet.cell(row=row, column=15, value="{}".format(variant.clinvar_data['hgvs']))
        tier3_fail_sheet.cell(row=row, column=16, value="{}".format(variant.clinvar_data['disease']))
        tier3_fail_sheet.cell(row=row, column=17,
                              value="{}".format(
                                  report_data['coverage'][variant.amplicon_data['amplicon']]['mean_coverage']))
        tier3_fail_sheet.cell(row=row, column=18,
                              value="{}".format(
                                  report_data['coverage'][variant.amplicon_data['amplicon']]['num_reads']))
        tier3_fail_sheet.cell(row=row, column=19, value="{}".format(variant.impact))
        tier3_fail_sheet.cell(row=row, column=20, value="{}".format(variant.severity))
        tier3_fail_sheet.cell(row=row, column=21, value="{}".format(variant.max_maf_all))
        tier3_fail_sheet.cell(row=row, column=22, value="{}".format(variant.min_depth))
        tier3_fail_sheet.cell(row=row, column=23, value="{}".format(variant.max_depth))
        tier3_fail_sheet.cell(row=row, column=24, value="{}".format(variant.chr))
        tier3_fail_sheet.cell(row=row, column=25, value="{}".format(variant.pos))
        tier3_fail_sheet.cell(row=row, column=26, value="{}".format(variant.end))
        tier3_fail_sheet.cell(row=row, column=27, value="{}".format(",".join(variant.rs_ids)))

        col = 28
        if 'mutect' in callers:
            tier3_fail_sheet.cell(row=row, column=col, value="{}".format(variant.mutect.get('AAF') or None))
            col += 1

        if 'vardict' in callers:
            tier3_fail_sheet.cell(row=row, column=col, value="{}".format(variant.vardict.get('AAF') or None))
            col += 1

        if 'freebayes' in callers:
            tier3_fail_sheet.cell(row=row, column=col, value="{}".format(variant.freebayes.get('AAF') or None))
            col += 1

        if 'scalpel' in callers:
            tier3_fail_sheet.cell(row=row, column=col, value="{}".format(variant.scalpel.get('AAF') or None))
            col += 1

        if 'platypus' in callers:
            tier3_fail_sheet.cell(row=row, column=col, value="{}".format(variant.platypus.get('AAF') or None))
            col += 1

        if 'pindel' in callers:
            tier3_fail_sheet.cell(row=row, column=col, value="{}".format(variant.pindel.get('AAF') or None))
            col += 1

        row += 1

    ####################################################################################################################

    tier4_fail_sheet.cell(row=1, column=1, value="Sample")
    tier4_fail_sheet.cell(row=1, column=2, value="Library")
    tier4_fail_sheet.cell(row=1, column=3, value="Gene")
    tier4_fail_sheet.cell(row=1, column=4, value="Amplicon")
    tier4_fail_sheet.cell(row=1, column=5, value="Ref")
    tier4_fail_sheet.cell(row=1, column=6, value="Alt")
    tier4_fail_sheet.cell(row=1, column=7, value="Codon")
    tier4_fail_sheet.cell(row=1, column=8, value="AA")
    tier4_fail_sheet.cell(row=1, column=9, value="Max Caller Somatic VAF")
    tier4_fail_sheet.cell(row=1, column=10, value="Callers")
    tier4_fail_sheet.cell(row=1, column=11, value="COSMIC IDs")
    tier4_fail_sheet.cell(row=1, column=12, value="Num COSMIC Samples")
    tier4_fail_sheet.cell(row=1, column=13, value="COSMIC AA")
    tier4_fail_sheet.cell(row=1, column=14, value="Clinvar Significance")
    tier4_fail_sheet.cell(row=1, column=15, value="Clinvar HGVS")
    tier4_fail_sheet.cell(row=1, column=16, value="Clinvar Disease")
    tier4_fail_sheet.cell(row=1, column=17, value="Coverage")
    tier4_fail_sheet.cell(row=1, column=18, value="Num Reads")
    tier4_fail_sheet.cell(row=1, column=19, value="Impact")
    tier4_fail_sheet.cell(row=1, column=20, value="Severity")
    tier4_fail_sheet.cell(row=1, column=21, value="Maximum Population AF")
    tier4_fail_sheet.cell(row=1, column=22, value="Min Caller Depth")
    tier4_fail_sheet.cell(row=1, column=23, value="Max Caller Depth")
    tier4_fail_sheet.cell(row=1, column=24, value="Chrom")
    tier4_fail_sheet.cell(row=1, column=25, value="Start")
    tier4_fail_sheet.cell(row=1, column=26, value="End")
    tier4_fail_sheet.cell(row=1, column=27, value="rsIDs")

    col = 28
    if 'mutect' in callers:
        tier4_fail_sheet.cell(row=1, column=col, value="MuTect_AF")
        col += 1

    if 'vardict' in callers:
        tier4_fail_sheet.cell(row=1, column=col, value="VarDict_AF")
        col += 1

    if 'freebayes' in callers:
        tier4_fail_sheet.cell(row=1, column=col, value="FreeBayes_AF")
        col += 1

    if 'scalpel' in callers:
        tier4_fail_sheet.cell(row=1, column=col, value="Scalpel_AF")
        col += 1

    if 'platypus' in callers:
        tier4_fail_sheet.cell(row=1, column=col, value="Platypus_AF")
        col += 1

    if 'pindel' in callers:
        tier4_fail_sheet.cell(row=1, column=col, value="Pindel_AF")
        col += 1

    row = 2
    for variant in report_data['variants']['tier4_fail_variants']:
        tier4_fail_sheet.cell(row=row, column=1, value="{}".format(variant.sample))
        tier4_fail_sheet.cell(row=row, column=2, value="{}".format(variant.library_name))
        tier4_fail_sheet.cell(row=row, column=3, value="{}".format(variant.gene))
        tier4_fail_sheet.cell(row=row, column=4, value="{}".format(variant.amplicon_data['amplicon']))
        tier4_fail_sheet.cell(row=row, column=5, value="{}".format(variant.ref))
        tier4_fail_sheet.cell(row=row, column=6, value="{}".format(variant.alt))
        tier4_fail_sheet.cell(row=row, column=7, value="{}".format(variant.codon_change))
        tier4_fail_sheet.cell(row=row, column=8, value="{}".format(variant.aa_change))
        tier4_fail_sheet.cell(row=row, column=9, value="{}".format(variant.max_som_aaf))
        tier4_fail_sheet.cell(row=row, column=10, value="{}".format(",".join(variant.callers) or None))
        tier4_fail_sheet.cell(row=row, column=11, value="{}".format(",".join(variant.cosmic_ids) or None))
        tier4_fail_sheet.cell(row=row, column=12, value="{}".format(variant.cosmic_data['num_samples']))
        tier4_fail_sheet.cell(row=row, column=13, value="{}".format(variant.cosmic_data['aa']))
        tier4_fail_sheet.cell(row=row, column=14, value="{}".format(variant.clinvar_data['significance']))
        tier4_fail_sheet.cell(row=row, column=15, value="{}".format(variant.clinvar_data['hgvs']))
        tier4_fail_sheet.cell(row=row, column=16, value="{}".format(variant.clinvar_data['disease']))
        tier4_fail_sheet.cell(row=row, column=17,
                              value="{}".format(
                                  report_data['coverage'][variant.amplicon_data['amplicon']]['mean_coverage']))
        tier4_fail_sheet.cell(row=row, column=18,
                              value="{}".format(
                                  report_data['coverage'][variant.amplicon_data['amplicon']]['num_reads']))
        tier4_fail_sheet.cell(row=row, column=19, value="{}".format(variant.impact))
        tier4_fail_sheet.cell(row=row, column=20, value="{}".format(variant.severity))
        tier4_fail_sheet.cell(row=row, column=21, value="{}".format(variant.max_maf_all))
        tier4_fail_sheet.cell(row=row, column=22, value="{}".format(variant.min_depth))
        tier4_fail_sheet.cell(row=row, column=23, value="{}".format(variant.max_depth))
        tier4_fail_sheet.cell(row=row, column=24, value="{}".format(variant.chr))
        tier4_fail_sheet.cell(row=row, column=25, value="{}".format(variant.pos))
        tier4_fail_sheet.cell(row=row, column=26, value="{}".format(variant.end))
        tier4_fail_sheet.cell(row=row, column=27, value="{}".format(",".join(variant.rs_ids)))

        col = 28
        if 'mutect' in callers:
            tier4_fail_sheet.cell(row=row, column=col, value="{}".format(variant.mutect.get('AAF') or None))
            col += 1

        if 'vardict' in callers:
            tier4_fail_sheet.cell(row=row, column=col, value="{}".format(variant.vardict.get('AAF') or None))
            col += 1

        if 'freebayes' in callers:
            tier4_fail_sheet.cell(row=row, column=col, value="{}".format(variant.freebayes.get('AAF') or None))
            col += 1

        if 'scalpel' in callers:
            tier4_fail_sheet.cell(row=row, column=col, value="{}".format(variant.scalpel.get('AAF') or None))
            col += 1

        if 'platypus' in callers:
            tier4_fail_sheet.cell(row=row, column=col, value="{}".format(variant.platypus.get('AAF') or None))
            col += 1

        if 'pindel' in callers:
            tier4_fail_sheet.cell(row=row, column=col, value="{}".format(variant.pindel.get('AAF') or None))
            col += 1

        row += 1

    ####################################################################################################################

    wb.save(report_name)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', '--samples_file', help="Input configuration file for samples")
    parser.add_argument('-c', '--configuration', help="Configuration file for various settings")
    parser.add_argument('-r', '--report', help="Root name for reports (per sample)", default='report')
    parser.add_argument('-a', '--address', help="IP Address for Cassandra connection", default='127.0.0.1')
    parser.add_argument('-u', '--username', help='Cassandra username for login', default=None)
    parser.add_argument('-d', '--min_depth', help='Minimum depth threshold for variant reporting', default=250.0)
    parser.add_argument('-t', '--min_somatic_var_freq', help='Minimum reportable somatic variant frequency',
                        default=0.01)
    parser.add_argument('-p', '--max_pop_freq', help='Maximum allowed population allele frequency', default=0.005)
    Job.Runner.addToilOptions(parser)
    args = parser.parse_args()
    args.logLevel = "INFO"

    sys.stdout.write("Parsing configuration data\n")
    config = configuration.configure_runtime(args.configuration)

    sys.stdout.write("Parsing sample data\n")
    libraries = configuration.configure_samples(args.samples_file, config)

    samples = configuration.merge_library_configs_samples(libraries)

    if args.username:
        password = getpass.getpass()
        auth_provider = PlainTextAuthProvider(username=args.username, password=password)
    else:
        auth_provider = None

    thresholds = {'min_saf': args.min_somatic_var_freq,
                  'max_maf': args.max_pop_freq,
                  'depth': args.min_depth}

    callers = ("mutect", "platypus", "vardict", "scalpel", "freebayes", "pindel")

    sys.stdout.write("Processing samples\n")
    root_job = Job.wrapJobFn(pipeline.spawn_batch_jobs, cores=1)

    for sample in samples:
        sample_job = Job.wrapJobFn(process_sample, config, sample, samples, [args.address], auth_provider,
                                   thresholds, callers, cores=1)

        root_job.addChild(sample_job)

    # Start workflow execution
    Job.Runner.startToil(root_job, args)
