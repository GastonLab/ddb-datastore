import utils
import numpy as np

from openpyxl import Workbook
from collections import defaultdict
from cassandra.cqlengine import connection
from variantstore import Variant
from variantstore import SampleVariant
from coveragestore import SampleCoverage


def get_variants(job, config, samples, sample, thresholds, amplicon_coverage, connection, addresses, authenticator):
    connection.setup(addresses, "variantstore", auth_provider=authenticator)
    filtered_sample_variants = list()
    report_data = dict()
    report_data['coverage'] = amplicon_coverage

    for library in samples[sample]:
        report_panel_path = "/mnt/shared-data/ddb-configs/disease_panels/{}/{}" \
                            "".format(samples[sample][library]['panel'], samples[sample][library]['report'])
        target_amplicons = utils.get_target_amplicons(report_panel_path)

        job.fileStore.logToMaster("{}: retrieving variants".format(library))
        variants = SampleVariant.objects.timeout(None).filter(
            SampleVariant.reference_genome == config['genome_version'],
            SampleVariant.sample == samples[sample][library]['sample_name'],
            SampleVariant.run_id == samples[sample][library]['run_id'],
            SampleVariant.library_name == samples[sample][library]['library_name'],
            SampleVariant.max_maf_all <= thresholds['max_maf']
        ).allow_filtering()

        num_var = variants.count()
        ordered = variants.order_by('library_name', 'chr', 'pos', 'ref', 'alt').limit(variants.count() + 1000)
        job.fileStore.logToMaster("{}: retrieved {} variants from database\n".format(library, num_var))

        filtered_job = job.addChildJobFn(classify_and_filter_variants, ordered, library, target_amplicons, config,
                                         thresholds, connection, addresses, authenticator)
        filtered_sample_variants.extend(filtered_job.rv())

    report_data['variants'] = filtered_sample_variants

    return report_data


def get_coverage_data(job, samples, sample, addresses, authenticator):
    job.fileStore.logToMaster("Retrieving coverage data for sample {}\n".format(sample))
    connection.setup(addresses, "coveragestore", auth_provider=authenticator)
    target_amplicon_coverage = dict()

    for library in samples[sample]:
        report_panel_path = "/mnt/shared-data/ddb-configs/disease_panels/{}/{}" \
                            "".format(samples[sample][library]['panel'], samples[sample][library]['report'])

        job.fileStore.logToMaster("{}: processing amplicons from file {}".format(library, report_panel_path))
        target_amplicons = utils.get_target_amplicons(report_panel_path)

        for amplicon in target_amplicons:
            coverage_data = SampleCoverage.objects.timeout(None).filter(
                SampleCoverage.sample == samples[sample][library]['sample_name'],
                SampleCoverage.amplicon == amplicon,
                SampleCoverage.run_id == samples[sample][library]['run_id'],
                SampleCoverage.library_name == samples[sample][library]['library_name'],
                SampleCoverage.program_name == "sambamba"
            )
            ordered_amplicons = coverage_data.order_by('amplicon', 'run_id').limit(coverage_data.count() + 1000)
            for result in ordered_amplicons:
                target_amplicon_coverage[amplicon]['num_reads'] = result.num_reads
                target_amplicon_coverage[amplicon]['mean_coverage'] = result.mean_coverage

    return target_amplicon_coverage


def classify_and_filter_variants(job, ordered_variants, library, target_amplicons, config, thresholds, connection,
                                 addresses, authenticator):
    job.fileStore.logToMaster("Filtering and classifying variants\n")
    connection.setup(addresses, "variantstore", auth_provider=authenticator)

    iterated = 0
    passing_variants = 0
    filtered_low_freq = 0
    filtered_low_depth = 0
    filtered_off_target = 0

    filtered_variant_data = defaultdict(list)
    off_target_amplicon_counts = defaultdict(int)

    job.fileStore.logToMaster("{}: classifying and filtering variants\n".format(library))

    for variant in ordered_variants:
        iterated += 1
        if variant.amplicon_data['amplicon'] is 'None':
            filtered_off_target += 1
            off_target_amplicon_counts[variant.amplicon_data['amplicon']] += 1
        else:
            amplicons = variant.amplicon_data['amplicon'].split(',')
            assignable = 0
            for amplicon in amplicons:
                if amplicon in target_amplicons:
                    assignable += 1
                    break
            if assignable:
                match_variants = Variant.objects.timeout(None).filter(
                    Variant.reference_genome == config['genome_version'],
                    Variant.chr == variant.chr,
                    Variant.pos == variant.pos,
                    Variant.ref == variant.ref,
                    Variant.alt == variant.alt
                ).allow_filtering()

                num_matches = match_variants.count()
                ordered_var = match_variants.order_by('sample', 'library_name', 'run_id').limit(num_matches + 1000)
                vafs = list()
                for var in ordered_var:
                    vaf = var.max_som_aaf
                    vafs.append(vaf)
                variant.vaf_median = np.median(vafs)
                variant.vaf_std_dev = np.std(vafs)

                # Putting in to Tier1 based on COSMIC
                if variant.cosmic_ids:
                    if variant.max_som_aaf < thresholds['min_saf']:
                        filtered_variant_data['tier1_fail_variants'].append(variant)
                        filtered_low_freq += 1
                    elif variant.max_depth < thresholds['depth']:
                        filtered_variant_data['tier1_fail_variants'].append(variant)
                        filtered_low_depth += 1
                    else:
                        filtered_variant_data['tier1_pass_variants'].append(variant)
                        passing_variants += 1
                    continue

                # Putting in to Tier1 based on ClinVar not being None or Benign
                if variant.clinvar_data['pathogenic'] != 'None':
                    if variant.clinvar_data['pathogenic'] != 'benign':
                        if variant.clinvar_data['pathogenic'] != 'likely-benign':
                            if variant.max_som_aaf < thresholds['min_saf']:
                                filtered_variant_data['tier1_fail_variants'].append(variant)
                                filtered_low_freq += 1
                            elif variant.max_depth < thresholds['depth']:
                                filtered_variant_data['tier1_fail_variants'].append(variant)
                                filtered_low_depth += 1
                            else:
                                filtered_variant_data['tier1_pass_variants'].append(variant)
                                passing_variants += 1
                            continue

                if variant.severity == 'MED' or variant.severity == 'HIGH':
                    if variant.max_som_aaf < thresholds['min_saf']:
                        filtered_variant_data['tier3_fail_variants'].append(variant)
                        filtered_low_freq += 1
                    elif variant.max_depth < thresholds['depth']:
                        filtered_variant_data['tier3_fail_variants'].append(variant)
                        filtered_low_depth += 1
                    else:
                        filtered_variant_data['tier3_pass_variants'].append(variant)
                        passing_variants += 1
                    continue
                else:
                    if variant.max_som_aaf < thresholds['min_saf']:
                        filtered_variant_data['tier4_fail_variants'].append(variant)
                        filtered_low_freq += 1
                    elif variant.max_depth < thresholds['depth']:
                        filtered_variant_data['tier4_fail_variants'].append(variant)
                        filtered_low_depth += 1
                    else:
                        filtered_variant_data['tier4_pass_variants'].append(variant)
                        passing_variants += 1
                    continue
            else:
                filtered_off_target += 1
                off_target_amplicon_counts[variant.amplicon_data['amplicon']] += 1

    job.fileStore.logToMaster("{}: iterated through {} variants\n".format(library, iterated))
    job.fileStore.logToMaster("{}: filtered {} off-target variants\n".format(library, filtered_off_target))
    job.fileStore.logToMaster("{}: filtered {} low-freq variants\n".format(library, filtered_low_freq))
    job.fileStore.logToMaster("{}: filtered {} low-depth variants\n".format(library, filtered_low_depth))

    job.fileStore.logToMaster("{}: passing {} tier 1 and 2 variants"
                              "\n".format(library, len(filtered_variant_data['tier1_pass_variants'])))
    job.fileStore.logToMaster("{}: passing {} tier3 variants"
                              "\n".format(library, len(filtered_variant_data['vus_pass_variants'])))
    job.fileStore.logToMaster("{}: passing {} tier 4 variants"
                              "\n".format(library, len(filtered_variant_data['tier4_pass_variants'])))

    return filtered_variant_data


def create_report(job, report_data, sample, samples, callers, thresholds):

    report_name = "{}.xlsx".format(sample)
    job.fileStore.logToMaster("{}: creating report {}\n".format(sample, report_name))

    wb = Workbook()
    coverage_sheet = wb.create_sheet(title="Coverage")
    tier1_sheet = wb.create_sheet(title="Tier1/2 Pass")
    tier3_sheet = wb.create_sheet(title="Tier3 Pass")
    tier4_sheet = wb.create_sheet(title="Tier4 Pass")
    tier1_fail_sheet = wb.create_sheet(title="Tier1/2 Fail")
    tier3_fail_sheet = wb.create_sheet(title="Tier3 Fail")
    tier4_fail_sheet = wb.create_sheet(title="Tier4 Fail")

    libraries = list()
    report_templates = list()
    run_id = ""
    for library in samples[sample]:
        libraries.append(samples[sample][library]['library_name'])
        report_templates.append(samples[sample][library]['report'])
        run_id = samples[sample][library]['run_id']
    lib_string = " | ".join(libraries)
    reports_string = " | ".join(report_templates)

    coverage_sheet.cell(row=0, column=0, value="Sample")
    coverage_sheet.cell(row=0, column=1, value="{}".format(sample))

    coverage_sheet.cell(row=1, column=0, value="Libraries")
    coverage_sheet.cell(row=1, column=1, value="{}".format(lib_string))

    coverage_sheet.cell(row=2, column=0, value="Run ID")
    coverage_sheet.cell(row=2, column=1, value="{}".format(run_id))

    coverage_sheet.cell(row=3, column=0, value="Reporting Templates")
    coverage_sheet.cell(row=3, column=1, value="{}".format(reports_string))

    coverage_sheet.cell(row=3, column=0, value="Minimum Reportable Somatic Allele Frequency")
    coverage_sheet.cell(row=3, column=1, value="{}".format(thresholds['min_saf']))

    coverage_sheet.cell(row=3, column=0, value="Minimum Amplicon Depth")
    coverage_sheet.cell(row=3, column=1, value="{}".format(thresholds['depth']))

    coverage_sheet.cell(row=4, column=0, value="Maximum Population Allele Frequency")
    coverage_sheet.cell(row=4, column=1, value="{}".format(thresholds['max_maf']))

    coverage_sheet.cell(row=5, column=0, value="Sample")
    coverage_sheet.cell(row=5, column=1, value="Library")
    coverage_sheet.cell(row=5, column=2, value="Amplicon")
    coverage_sheet.cell(row=5, column=3, value="Num Reads")
    coverage_sheet.cell(row=5, column=4, value="Coverage")

    row_num = 6
    for amplicon in report_data['coverage']:
        coverage_sheet.cell(row=row_num, column=0, value="{}".format(report_data['coverage'][amplicon].sample))
        coverage_sheet.cell(row=row_num, column=1, value="{}".format(report_data['coverage'][amplicon].library_name))
        coverage_sheet.cell(row=row_num, column=2, value="{}".format(report_data['coverage'][amplicon].amplicon))
        coverage_sheet.cell(row=row_num, column=3, value="{}".format(report_data['coverage'][amplicon].num_reads))
        coverage_sheet.cell(row=row_num, column=4, value="{}".format(report_data['coverage'][amplicon].mean_coverage))

        row_num += 1

    ####################################################################################################################

    tier1_sheet.cell(row=0, column=0, value="Sample")
    tier1_sheet.cell(row=0, column=1, value="Library")
    tier1_sheet.cell(row=0, column=2, value="Gene")
    tier1_sheet.cell(row=0, column=3, value="Amplicon")
    tier1_sheet.cell(row=0, column=4, value="Ref")
    tier1_sheet.cell(row=0, column=5, value="Alt")
    tier1_sheet.cell(row=0, column=6, value="Codon")
    tier1_sheet.cell(row=0, column=7, value="AA")
    tier1_sheet.cell(row=0, column=8, value="Max Caller Somatic VAF")
    tier1_sheet.cell(row=0, column=9, value="Callers")
    tier1_sheet.cell(row=0, column=10, value="COSMIC IDs")
    tier1_sheet.cell(row=0, column=11, value="Num COSMIC Samples")
    tier1_sheet.cell(row=0, column=12, value="COSMIC AA")
    tier1_sheet.cell(row=0, column=13, value="Clinvar Significance")
    tier1_sheet.cell(row=0, column=14, value="Clinvar HGVS")
    tier1_sheet.cell(row=0, column=15, value="Clinvar Disease")
    tier1_sheet.cell(row=0, column=16, value="Coverage")
    tier1_sheet.cell(row=0, column=17, value="Num Reads")
    tier1_sheet.cell(row=0, column=18, value="Impact")
    tier1_sheet.cell(row=0, column=19, value="Severity")
    tier1_sheet.cell(row=0, column=20, value="Maximum Population AF")
    tier1_sheet.cell(row=0, column=21, value="Min Caller Depth")
    tier1_sheet.cell(row=0, column=22, value="Max Caller Depth")
    tier1_sheet.cell(row=0, column=23, value="Chrom")
    tier1_sheet.cell(row=0, column=24, value="Start")
    tier1_sheet.cell(row=0, column=25, value="End")
    tier1_sheet.cell(row=0, column=26, value="rsIDs")

    col = 27
    if 'mutect' in callers:
        tier1_sheet.cell(row=0, column=col, value="MuTect_AF")
        col += 1

    if 'vardict' in callers:
        tier1_sheet.cell(row=0, column=col, value="VarDict_AF")
        col += 1

    if 'freebayes' in callers:
        tier1_sheet.cell(row=0, column=col, value="FreeBayes_AF")
        col += 1

    if 'scalpel' in callers:
        tier1_sheet.cell(row=0, column=col, value="Scalpel_AF")
        col += 1

    if 'platypus' in callers:
        tier1_sheet.cell(row=0, column=col, value="Platypus_AF")
        col += 1

    if 'pindel' in callers:
        tier1_sheet.cell(row=0, column=col, value="Pindel_AF")
        col += 1

    row = 1
    for variant in report_data['']:
        tier1_sheet.cell(row=row, column=0, value="{}".format(variant.sample))
        tier1_sheet.cell(row=row, column=1, value="{}".format(variant.library_name))
        tier1_sheet.cell(row=row, column=2, value="{}".format(variant.gene))
        tier1_sheet.cell(row=row, column=3, value="{}".format(variant.amplicon_data['amplicon']))
        tier1_sheet.cell(row=row, column=4, value="{}".format(variant.ref))
        tier1_sheet.cell(row=row, column=5, value="{}".format(variant.alt))
        tier1_sheet.cell(row=row, column=6, value="{}".format(variant.codon_change))
        tier1_sheet.cell(row=row, column=7, value="{}".format(variant.aa_change))
        tier1_sheet.cell(row=row, column=8, value="{}".format(variant.max_som_aaf))
        tier1_sheet.cell(row=row, column=9, value="{}".format(",".join(variant.callers) or None))
        tier1_sheet.cell(row=row, column=10, value="{}".format(",".join(variant.cosmic_ids) or None))
        tier1_sheet.cell(row=row, column=11, value="{}".format(variant.cosmic_data['num_samples']))
        tier1_sheet.cell(row=row, column=12, value="{}".format(variant.cosmic_data['aa']))
        tier1_sheet.cell(row=row, column=13, value="{}".format(variant.clinvar_data['significance']))
        tier1_sheet.cell(row=row, column=14, value="{}".format(variant.clinvar_data['hgvs']))
        tier1_sheet.cell(row=row, column=15, value="{}".format(variant.clinvar_data['disease']))
        tier1_sheet.cell(row=row, column=16, value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['mean_coverage']))
        tier1_sheet.cell(row=row, column=17, value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['num_reads']))
        tier1_sheet.cell(row=row, column=18, value="{}".format(variant.impact))
        tier1_sheet.cell(row=row, column=19, value="{}".format(variant.severity))
        tier1_sheet.cell(row=row, column=20, value="{}".format(variant.max_maf_all))
        tier1_sheet.cell(row=row, column=21, value="{}".format(variant.min_depth))
        tier1_sheet.cell(row=row, column=22, value="{}".format(variant.max_depth))
        tier1_sheet.cell(row=row, column=23, value="{}".format(variant.chr))
        tier1_sheet.cell(row=row, column=24, value="{}".format(variant.pos))
        tier1_sheet.cell(row=row, column=25, value="{}".format(variant.end))
        tier1_sheet.cell(row=row, column=26, value="{}".format(",".join(variant.rs_ids)))

        col = 27
        if 'mutect' in callers:
            tier1_sheet.cell(row=row, column=col, value="{}".format(variant.mutect.get('AAF') or None))
            col += 1

        if 'vardict' in callers:
            tier1_sheet.cell(row=row, column=col, value="{}".format(variant.vardict.get('AAF') or None))
            col += 1

        if 'freebayes' in callers:
            tier1_sheet.cell(row=row, column=col, value="{}".format(variant.freebayes.get('AAF') or None))
            col += 1

        if 'scalpel' in callers:
            tier1_sheet.cell(row=row, column=col, value="{}".format(variant.scalpel.get('AAF') or None))
            col += 1

        if 'platypus' in callers:
            tier1_sheet.cell(row=row, column=col, value="{}".format(variant.platypus.get('AAF') or None))
            col += 1

        if 'pindel' in callers:
            tier1_sheet.cell(row=row, column=col, value="{}".format(variant.pindel.get('AAF') or None))
            col += 1

        row += 1

    ####################################################################################################################

    tier3_sheet.cell(row=0, column=0, value="Sample")
    tier3_sheet.cell(row=0, column=1, value="Library")
    tier3_sheet.cell(row=0, column=2, value="Gene")
    tier3_sheet.cell(row=0, column=3, value="Amplicon")
    tier3_sheet.cell(row=0, column=4, value="Ref")
    tier3_sheet.cell(row=0, column=5, value="Alt")
    tier3_sheet.cell(row=0, column=6, value="Codon")
    tier3_sheet.cell(row=0, column=7, value="AA")
    tier3_sheet.cell(row=0, column=8, value="Max Caller Somatic VAF")
    tier3_sheet.cell(row=0, column=9, value="Callers")
    tier3_sheet.cell(row=0, column=10, value="COSMIC IDs")
    tier3_sheet.cell(row=0, column=11, value="Num COSMIC Samples")
    tier3_sheet.cell(row=0, column=12, value="COSMIC AA")
    tier3_sheet.cell(row=0, column=13, value="Clinvar Significance")
    tier3_sheet.cell(row=0, column=14, value="Clinvar HGVS")
    tier3_sheet.cell(row=0, column=15, value="Clinvar Disease")
    tier3_sheet.cell(row=0, column=16, value="Coverage")
    tier3_sheet.cell(row=0, column=17, value="Num Reads")
    tier3_sheet.cell(row=0, column=18, value="Impact")
    tier3_sheet.cell(row=0, column=19, value="Severity")
    tier3_sheet.cell(row=0, column=20, value="Maximum Population AF")
    tier3_sheet.cell(row=0, column=21, value="Min Caller Depth")
    tier3_sheet.cell(row=0, column=22, value="Max Caller Depth")
    tier3_sheet.cell(row=0, column=23, value="Chrom")
    tier3_sheet.cell(row=0, column=24, value="Start")
    tier3_sheet.cell(row=0, column=25, value="End")
    tier3_sheet.cell(row=0, column=26, value="rsIDs")

    col = 27
    if 'mutect' in callers:
        tier3_sheet.cell(row=0, column=col, value="MuTect_AF")
        col += 1

    if 'vardict' in callers:
        tier3_sheet.cell(row=0, column=col, value="VarDict_AF")
        col += 1

    if 'freebayes' in callers:
        tier3_sheet.cell(row=0, column=col, value="FreeBayes_AF")
        col += 1

    if 'scalpel' in callers:
        tier3_sheet.cell(row=0, column=col, value="Scalpel_AF")
        col += 1

    if 'platypus' in callers:
        tier3_sheet.cell(row=0, column=col, value="Platypus_AF")
        col += 1

    if 'pindel' in callers:
        tier3_sheet.cell(row=0, column=col, value="Pindel_AF")
        col += 1

    row = 1
    for variant in report_data['']:
        tier3_sheet.cell(row=row, column=0, value="{}".format(variant.sample))
        tier3_sheet.cell(row=row, column=1, value="{}".format(variant.library_name))
        tier3_sheet.cell(row=row, column=2, value="{}".format(variant.gene))
        tier3_sheet.cell(row=row, column=3, value="{}".format(variant.amplicon_data['amplicon']))
        tier3_sheet.cell(row=row, column=4, value="{}".format(variant.ref))
        tier3_sheet.cell(row=row, column=5, value="{}".format(variant.alt))
        tier3_sheet.cell(row=row, column=6, value="{}".format(variant.codon_change))
        tier3_sheet.cell(row=row, column=7, value="{}".format(variant.aa_change))
        tier3_sheet.cell(row=row, column=8, value="{}".format(variant.max_som_aaf))
        tier3_sheet.cell(row=row, column=9, value="{}".format(",".join(variant.callers) or None))
        tier3_sheet.cell(row=row, column=10, value="{}".format(",".join(variant.cosmic_ids) or None))
        tier3_sheet.cell(row=row, column=11, value="{}".format(variant.cosmic_data['num_samples']))
        tier3_sheet.cell(row=row, column=12, value="{}".format(variant.cosmic_data['aa']))
        tier3_sheet.cell(row=row, column=13, value="{}".format(variant.clinvar_data['significance']))
        tier3_sheet.cell(row=row, column=14, value="{}".format(variant.clinvar_data['hgvs']))
        tier3_sheet.cell(row=row, column=15, value="{}".format(variant.clinvar_data['disease']))
        tier3_sheet.cell(row=row, column=16,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['mean_coverage']))
        tier3_sheet.cell(row=row, column=17,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['num_reads']))
        tier3_sheet.cell(row=row, column=18, value="{}".format(variant.impact))
        tier3_sheet.cell(row=row, column=19, value="{}".format(variant.severity))
        tier3_sheet.cell(row=row, column=20, value="{}".format(variant.max_maf_all))
        tier3_sheet.cell(row=row, column=21, value="{}".format(variant.min_depth))
        tier3_sheet.cell(row=row, column=22, value="{}".format(variant.max_depth))
        tier3_sheet.cell(row=row, column=23, value="{}".format(variant.chr))
        tier3_sheet.cell(row=row, column=24, value="{}".format(variant.pos))
        tier3_sheet.cell(row=row, column=25, value="{}".format(variant.end))
        tier3_sheet.cell(row=row, column=26, value="{}".format(",".join(variant.rs_ids)))

        col = 27
        if 'mutect' in callers:
            tier3_sheet.cell(row=row, column=col, value="{}".format(variant.mutect.get('AAF') or None))
            col += 1

        if 'vardict' in callers:
            tier3_sheet.cell(row=row, column=col, value="{}".format(variant.vardict.get('AAF') or None))
            col += 1

        if 'freebayes' in callers:
            tier3_sheet.cell(row=row, column=col, value="{}".format(variant.freebayes.get('AAF') or None))
            col += 1

        if 'scalpel' in callers:
            tier3_sheet.cell(row=row, column=col, value="{}".format(variant.scalpel.get('AAF') or None))
            col += 1

        if 'platypus' in callers:
            tier3_sheet.cell(row=row, column=col, value="{}".format(variant.platypus.get('AAF') or None))
            col += 1

        if 'pindel' in callers:
            tier3_sheet.cell(row=row, column=col, value="{}".format(variant.pindel.get('AAF') or None))
            col += 1

        row += 1

    ####################################################################################################################

    tier4_sheet.cell(row=0, column=0, value="Sample")
    tier4_sheet.cell(row=0, column=1, value="Library")
    tier4_sheet.cell(row=0, column=2, value="Gene")
    tier4_sheet.cell(row=0, column=3, value="Amplicon")
    tier4_sheet.cell(row=0, column=4, value="Ref")
    tier4_sheet.cell(row=0, column=5, value="Alt")
    tier4_sheet.cell(row=0, column=6, value="Codon")
    tier4_sheet.cell(row=0, column=7, value="AA")
    tier4_sheet.cell(row=0, column=8, value="Max Caller Somatic VAF")
    tier4_sheet.cell(row=0, column=9, value="Callers")
    tier4_sheet.cell(row=0, column=10, value="COSMIC IDs")
    tier4_sheet.cell(row=0, column=11, value="Num COSMIC Samples")
    tier4_sheet.cell(row=0, column=12, value="COSMIC AA")
    tier4_sheet.cell(row=0, column=13, value="Clinvar Significance")
    tier4_sheet.cell(row=0, column=14, value="Clinvar HGVS")
    tier4_sheet.cell(row=0, column=15, value="Clinvar Disease")
    tier4_sheet.cell(row=0, column=16, value="Coverage")
    tier4_sheet.cell(row=0, column=17, value="Num Reads")
    tier4_sheet.cell(row=0, column=18, value="Impact")
    tier4_sheet.cell(row=0, column=19, value="Severity")
    tier4_sheet.cell(row=0, column=20, value="Maximum Population AF")
    tier4_sheet.cell(row=0, column=21, value="Min Caller Depth")
    tier4_sheet.cell(row=0, column=22, value="Max Caller Depth")
    tier4_sheet.cell(row=0, column=23, value="Chrom")
    tier4_sheet.cell(row=0, column=24, value="Start")
    tier4_sheet.cell(row=0, column=25, value="End")
    tier4_sheet.cell(row=0, column=26, value="rsIDs")

    col = 27
    if 'mutect' in callers:
        tier4_sheet.cell(row=0, column=col, value="MuTect_AF")
        col += 1

    if 'vardict' in callers:
        tier4_sheet.cell(row=0, column=col, value="VarDict_AF")
        col += 1

    if 'freebayes' in callers:
        tier4_sheet.cell(row=0, column=col, value="FreeBayes_AF")
        col += 1

    if 'scalpel' in callers:
        tier4_sheet.cell(row=0, column=col, value="Scalpel_AF")
        col += 1

    if 'platypus' in callers:
        tier4_sheet.cell(row=0, column=col, value="Platypus_AF")
        col += 1

    if 'pindel' in callers:
        tier4_sheet.cell(row=0, column=col, value="Pindel_AF")
        col += 1

    row = 1
    for variant in report_data['']:
        tier4_sheet.cell(row=row, column=0, value="{}".format(variant.sample))
        tier4_sheet.cell(row=row, column=1, value="{}".format(variant.library_name))
        tier4_sheet.cell(row=row, column=2, value="{}".format(variant.gene))
        tier4_sheet.cell(row=row, column=3, value="{}".format(variant.amplicon_data['amplicon']))
        tier4_sheet.cell(row=row, column=4, value="{}".format(variant.ref))
        tier4_sheet.cell(row=row, column=5, value="{}".format(variant.alt))
        tier4_sheet.cell(row=row, column=6, value="{}".format(variant.codon_change))
        tier4_sheet.cell(row=row, column=7, value="{}".format(variant.aa_change))
        tier4_sheet.cell(row=row, column=8, value="{}".format(variant.max_som_aaf))
        tier4_sheet.cell(row=row, column=9, value="{}".format(",".join(variant.callers) or None))
        tier4_sheet.cell(row=row, column=10, value="{}".format(",".join(variant.cosmic_ids) or None))
        tier4_sheet.cell(row=row, column=11, value="{}".format(variant.cosmic_data['num_samples']))
        tier4_sheet.cell(row=row, column=12, value="{}".format(variant.cosmic_data['aa']))
        tier4_sheet.cell(row=row, column=13, value="{}".format(variant.clinvar_data['significance']))
        tier4_sheet.cell(row=row, column=14, value="{}".format(variant.clinvar_data['hgvs']))
        tier4_sheet.cell(row=row, column=15, value="{}".format(variant.clinvar_data['disease']))
        tier4_sheet.cell(row=row, column=16,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['mean_coverage']))
        tier4_sheet.cell(row=row, column=17,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['num_reads']))
        tier4_sheet.cell(row=row, column=18, value="{}".format(variant.impact))
        tier4_sheet.cell(row=row, column=19, value="{}".format(variant.severity))
        tier4_sheet.cell(row=row, column=20, value="{}".format(variant.max_maf_all))
        tier4_sheet.cell(row=row, column=21, value="{}".format(variant.min_depth))
        tier4_sheet.cell(row=row, column=22, value="{}".format(variant.max_depth))
        tier4_sheet.cell(row=row, column=23, value="{}".format(variant.chr))
        tier4_sheet.cell(row=row, column=24, value="{}".format(variant.pos))
        tier4_sheet.cell(row=row, column=25, value="{}".format(variant.end))
        tier4_sheet.cell(row=row, column=26, value="{}".format(",".join(variant.rs_ids)))

        col = 27
        if 'mutect' in callers:
            tier4_sheet.cell(row=row, column=col, value="{}".format(variant.mutect.get('AAF') or None))
            col += 1

        if 'vardict' in callers:
            tier4_sheet.cell(row=row, column=col, value="{}".format(variant.vardict.get('AAF') or None))
            col += 1

        if 'freebayes' in callers:
            tier4_sheet.cell(row=row, column=col, value="{}".format(variant.freebayes.get('AAF') or None))
            col += 1

        if 'scalpel' in callers:
            tier4_sheet.cell(row=row, column=col, value="{}".format(variant.scalpel.get('AAF') or None))
            col += 1

        if 'platypus' in callers:
            tier4_sheet.cell(row=row, column=col, value="{}".format(variant.platypus.get('AAF') or None))
            col += 1

        if 'pindel' in callers:
            tier4_sheet.cell(row=row, column=col, value="{}".format(variant.pindel.get('AAF') or None))
            col += 1

        row += 1

    ####################################################################################################################

    tier1_fail_sheet.cell(row=0, column=0, value="Sample")
    tier1_fail_sheet.cell(row=0, column=1, value="Library")
    tier1_fail_sheet.cell(row=0, column=2, value="Gene")
    tier1_fail_sheet.cell(row=0, column=3, value="Amplicon")
    tier1_fail_sheet.cell(row=0, column=4, value="Ref")
    tier1_fail_sheet.cell(row=0, column=5, value="Alt")
    tier1_fail_sheet.cell(row=0, column=6, value="Codon")
    tier1_fail_sheet.cell(row=0, column=7, value="AA")
    tier1_fail_sheet.cell(row=0, column=8, value="Max Caller Somatic VAF")
    tier1_fail_sheet.cell(row=0, column=9, value="Callers")
    tier1_fail_sheet.cell(row=0, column=10, value="COSMIC IDs")
    tier1_fail_sheet.cell(row=0, column=11, value="Num COSMIC Samples")
    tier1_fail_sheet.cell(row=0, column=12, value="COSMIC AA")
    tier1_fail_sheet.cell(row=0, column=13, value="Clinvar Significance")
    tier1_fail_sheet.cell(row=0, column=14, value="Clinvar HGVS")
    tier1_fail_sheet.cell(row=0, column=15, value="Clinvar Disease")
    tier1_fail_sheet.cell(row=0, column=16, value="Coverage")
    tier1_fail_sheet.cell(row=0, column=17, value="Num Reads")
    tier1_fail_sheet.cell(row=0, column=18, value="Impact")
    tier1_fail_sheet.cell(row=0, column=19, value="Severity")
    tier1_fail_sheet.cell(row=0, column=20, value="Maximum Population AF")
    tier1_fail_sheet.cell(row=0, column=21, value="Min Caller Depth")
    tier1_fail_sheet.cell(row=0, column=22, value="Max Caller Depth")
    tier1_fail_sheet.cell(row=0, column=23, value="Chrom")
    tier1_fail_sheet.cell(row=0, column=24, value="Start")
    tier1_fail_sheet.cell(row=0, column=25, value="End")
    tier1_fail_sheet.cell(row=0, column=26, value="rsIDs")

    col = 27
    if 'mutect' in callers:
        tier1_fail_sheet.cell(row=0, column=col, value="MuTect_AF")
        col += 1

    if 'vardict' in callers:
        tier1_fail_sheet.cell(row=0, column=col, value="VarDict_AF")
        col += 1

    if 'freebayes' in callers:
        tier1_fail_sheet.cell(row=0, column=col, value="FreeBayes_AF")
        col += 1

    if 'scalpel' in callers:
        tier1_fail_sheet.cell(row=0, column=col, value="Scalpel_AF")
        col += 1

    if 'platypus' in callers:
        tier1_fail_sheet.cell(row=0, column=col, value="Platypus_AF")
        col += 1

    if 'pindel' in callers:
        tier1_fail_sheet.cell(row=0, column=col, value="Pindel_AF")
        col += 1

    row = 1
    for variant in report_data['']:
        tier1_fail_sheet.cell(row=row, column=0, value="{}".format(variant.sample))
        tier1_fail_sheet.cell(row=row, column=1, value="{}".format(variant.library_name))
        tier1_fail_sheet.cell(row=row, column=2, value="{}".format(variant.gene))
        tier1_fail_sheet.cell(row=row, column=3, value="{}".format(variant.amplicon_data['amplicon']))
        tier1_fail_sheet.cell(row=row, column=4, value="{}".format(variant.ref))
        tier1_fail_sheet.cell(row=row, column=5, value="{}".format(variant.alt))
        tier1_fail_sheet.cell(row=row, column=6, value="{}".format(variant.codon_change))
        tier1_fail_sheet.cell(row=row, column=7, value="{}".format(variant.aa_change))
        tier1_fail_sheet.cell(row=row, column=8, value="{}".format(variant.max_som_aaf))
        tier1_fail_sheet.cell(row=row, column=9, value="{}".format(",".join(variant.callers) or None))
        tier1_fail_sheet.cell(row=row, column=10, value="{}".format(",".join(variant.cosmic_ids) or None))
        tier1_fail_sheet.cell(row=row, column=11, value="{}".format(variant.cosmic_data['num_samples']))
        tier1_fail_sheet.cell(row=row, column=12, value="{}".format(variant.cosmic_data['aa']))
        tier1_fail_sheet.cell(row=row, column=13, value="{}".format(variant.clinvar_data['significance']))
        tier1_fail_sheet.cell(row=row, column=14, value="{}".format(variant.clinvar_data['hgvs']))
        tier1_fail_sheet.cell(row=row, column=15, value="{}".format(variant.clinvar_data['disease']))
        tier1_fail_sheet.cell(row=row, column=16,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['mean_coverage']))
        tier1_fail_sheet.cell(row=row, column=17,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['num_reads']))
        tier1_fail_sheet.cell(row=row, column=18, value="{}".format(variant.impact))
        tier1_fail_sheet.cell(row=row, column=19, value="{}".format(variant.severity))
        tier1_fail_sheet.cell(row=row, column=20, value="{}".format(variant.max_maf_all))
        tier1_fail_sheet.cell(row=row, column=21, value="{}".format(variant.min_depth))
        tier1_fail_sheet.cell(row=row, column=22, value="{}".format(variant.max_depth))
        tier1_fail_sheet.cell(row=row, column=23, value="{}".format(variant.chr))
        tier1_fail_sheet.cell(row=row, column=24, value="{}".format(variant.pos))
        tier1_fail_sheet.cell(row=row, column=25, value="{}".format(variant.end))
        tier1_fail_sheet.cell(row=row, column=26, value="{}".format(",".join(variant.rs_ids)))

        col = 27
        if 'mutect' in callers:
            tier1_fail_sheet.cell(row=row, column=col, value="{}".format(variant.mutect.get('AAF') or None))
            col += 1

        if 'vardict' in callers:
            tier1_fail_sheet.cell(row=row, column=col, value="{}".format(variant.vardict.get('AAF') or None))
            col += 1

        if 'freebayes' in callers:
            tier1_fail_sheet.cell(row=row, column=col, value="{}".format(variant.freebayes.get('AAF') or None))
            col += 1

        if 'scalpel' in callers:
            tier1_fail_sheet.cell(row=row, column=col, value="{}".format(variant.scalpel.get('AAF') or None))
            col += 1

        if 'platypus' in callers:
            tier1_fail_sheet.cell(row=row, column=col, value="{}".format(variant.platypus.get('AAF') or None))
            col += 1

        if 'pindel' in callers:
            tier1_fail_sheet.cell(row=row, column=col, value="{}".format(variant.pindel.get('AAF') or None))
            col += 1

        row += 1

    ####################################################################################################################

    tier3_fail_sheet.cell(row=0, column=0, value="Sample")
    tier3_fail_sheet.cell(row=0, column=1, value="Library")
    tier3_fail_sheet.cell(row=0, column=2, value="Gene")
    tier3_fail_sheet.cell(row=0, column=3, value="Amplicon")
    tier3_fail_sheet.cell(row=0, column=4, value="Ref")
    tier3_fail_sheet.cell(row=0, column=5, value="Alt")
    tier3_fail_sheet.cell(row=0, column=6, value="Codon")
    tier3_fail_sheet.cell(row=0, column=7, value="AA")
    tier3_fail_sheet.cell(row=0, column=8, value="Max Caller Somatic VAF")
    tier3_fail_sheet.cell(row=0, column=9, value="Callers")
    tier3_fail_sheet.cell(row=0, column=10, value="COSMIC IDs")
    tier3_fail_sheet.cell(row=0, column=11, value="Num COSMIC Samples")
    tier3_fail_sheet.cell(row=0, column=12, value="COSMIC AA")
    tier3_fail_sheet.cell(row=0, column=13, value="Clinvar Significance")
    tier3_fail_sheet.cell(row=0, column=14, value="Clinvar HGVS")
    tier3_fail_sheet.cell(row=0, column=15, value="Clinvar Disease")
    tier3_fail_sheet.cell(row=0, column=16, value="Coverage")
    tier3_fail_sheet.cell(row=0, column=17, value="Num Reads")
    tier3_fail_sheet.cell(row=0, column=18, value="Impact")
    tier3_fail_sheet.cell(row=0, column=19, value="Severity")
    tier3_fail_sheet.cell(row=0, column=20, value="Maximum Population AF")
    tier3_fail_sheet.cell(row=0, column=21, value="Min Caller Depth")
    tier3_fail_sheet.cell(row=0, column=22, value="Max Caller Depth")
    tier3_fail_sheet.cell(row=0, column=23, value="Chrom")
    tier3_fail_sheet.cell(row=0, column=24, value="Start")
    tier3_fail_sheet.cell(row=0, column=25, value="End")
    tier3_fail_sheet.cell(row=0, column=26, value="rsIDs")

    col = 27
    if 'mutect' in callers:
        tier3_fail_sheet.cell(row=0, column=col, value="MuTect_AF")
        col += 1

    if 'vardict' in callers:
        tier3_fail_sheet.cell(row=0, column=col, value="VarDict_AF")
        col += 1

    if 'freebayes' in callers:
        tier3_fail_sheet.cell(row=0, column=col, value="FreeBayes_AF")
        col += 1

    if 'scalpel' in callers:
        tier3_fail_sheet.cell(row=0, column=col, value="Scalpel_AF")
        col += 1

    if 'platypus' in callers:
        tier3_fail_sheet.cell(row=0, column=col, value="Platypus_AF")
        col += 1

    if 'pindel' in callers:
        tier3_fail_sheet.cell(row=0, column=col, value="Pindel_AF")
        col += 1

    row = 1
    for variant in report_data['']:
        tier3_fail_sheet.cell(row=row, column=0, value="{}".format(variant.sample))
        tier3_fail_sheet.cell(row=row, column=1, value="{}".format(variant.library_name))
        tier3_fail_sheet.cell(row=row, column=2, value="{}".format(variant.gene))
        tier3_fail_sheet.cell(row=row, column=3, value="{}".format(variant.amplicon_data['amplicon']))
        tier3_fail_sheet.cell(row=row, column=4, value="{}".format(variant.ref))
        tier3_fail_sheet.cell(row=row, column=5, value="{}".format(variant.alt))
        tier3_fail_sheet.cell(row=row, column=6, value="{}".format(variant.codon_change))
        tier3_fail_sheet.cell(row=row, column=7, value="{}".format(variant.aa_change))
        tier3_fail_sheet.cell(row=row, column=8, value="{}".format(variant.max_som_aaf))
        tier3_fail_sheet.cell(row=row, column=9, value="{}".format(",".join(variant.callers) or None))
        tier3_fail_sheet.cell(row=row, column=10, value="{}".format(",".join(variant.cosmic_ids) or None))
        tier3_fail_sheet.cell(row=row, column=11, value="{}".format(variant.cosmic_data['num_samples']))
        tier3_fail_sheet.cell(row=row, column=12, value="{}".format(variant.cosmic_data['aa']))
        tier3_fail_sheet.cell(row=row, column=13, value="{}".format(variant.clinvar_data['significance']))
        tier3_fail_sheet.cell(row=row, column=14, value="{}".format(variant.clinvar_data['hgvs']))
        tier3_fail_sheet.cell(row=row, column=15, value="{}".format(variant.clinvar_data['disease']))
        tier3_fail_sheet.cell(row=row, column=16,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['mean_coverage']))
        tier3_fail_sheet.cell(row=row, column=17,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['num_reads']))
        tier3_fail_sheet.cell(row=row, column=18, value="{}".format(variant.impact))
        tier3_fail_sheet.cell(row=row, column=19, value="{}".format(variant.severity))
        tier3_fail_sheet.cell(row=row, column=20, value="{}".format(variant.max_maf_all))
        tier3_fail_sheet.cell(row=row, column=21, value="{}".format(variant.min_depth))
        tier3_fail_sheet.cell(row=row, column=22, value="{}".format(variant.max_depth))
        tier3_fail_sheet.cell(row=row, column=23, value="{}".format(variant.chr))
        tier3_fail_sheet.cell(row=row, column=24, value="{}".format(variant.pos))
        tier3_fail_sheet.cell(row=row, column=25, value="{}".format(variant.end))
        tier3_fail_sheet.cell(row=row, column=26, value="{}".format(",".join(variant.rs_ids)))

        col = 27
        if 'mutect' in callers:
            tier3_fail_sheet.cell(row=row, column=col, value="{}".format(variant.mutect.get('AAF') or None))
            col += 1

        if 'vardict' in callers:
            tier3_fail_sheet.cell(row=row, column=col, value="{}".format(variant.vardict.get('AAF') or None))
            col += 1

        if 'freebayes' in callers:
            tier3_fail_sheet.cell(row=row, column=col, value="{}".format(variant.freebayes.get('AAF') or None))
            col += 1

        if 'scalpel' in callers:
            tier3_fail_sheet.cell(row=row, column=col, value="{}".format(variant.scalpel.get('AAF') or None))
            col += 1

        if 'platypus' in callers:
            tier3_fail_sheet.cell(row=row, column=col, value="{}".format(variant.platypus.get('AAF') or None))
            col += 1

        if 'pindel' in callers:
            tier3_fail_sheet.cell(row=row, column=col, value="{}".format(variant.pindel.get('AAF') or None))
            col += 1

        row += 1

    ####################################################################################################################

    tier4_fail_sheet.cell(row=0, column=0, value="Sample")
    tier4_fail_sheet.cell(row=0, column=1, value="Library")
    tier4_fail_sheet.cell(row=0, column=2, value="Gene")
    tier4_fail_sheet.cell(row=0, column=3, value="Amplicon")
    tier4_fail_sheet.cell(row=0, column=4, value="Ref")
    tier4_fail_sheet.cell(row=0, column=5, value="Alt")
    tier4_fail_sheet.cell(row=0, column=6, value="Codon")
    tier4_fail_sheet.cell(row=0, column=7, value="AA")
    tier4_fail_sheet.cell(row=0, column=8, value="Max Caller Somatic VAF")
    tier4_fail_sheet.cell(row=0, column=9, value="Callers")
    tier4_fail_sheet.cell(row=0, column=10, value="COSMIC IDs")
    tier4_fail_sheet.cell(row=0, column=11, value="Num COSMIC Samples")
    tier4_fail_sheet.cell(row=0, column=12, value="COSMIC AA")
    tier4_fail_sheet.cell(row=0, column=13, value="Clinvar Significance")
    tier4_fail_sheet.cell(row=0, column=14, value="Clinvar HGVS")
    tier4_fail_sheet.cell(row=0, column=15, value="Clinvar Disease")
    tier4_fail_sheet.cell(row=0, column=16, value="Coverage")
    tier4_fail_sheet.cell(row=0, column=17, value="Num Reads")
    tier4_fail_sheet.cell(row=0, column=18, value="Impact")
    tier4_fail_sheet.cell(row=0, column=19, value="Severity")
    tier4_fail_sheet.cell(row=0, column=20, value="Maximum Population AF")
    tier4_fail_sheet.cell(row=0, column=21, value="Min Caller Depth")
    tier4_fail_sheet.cell(row=0, column=22, value="Max Caller Depth")
    tier4_fail_sheet.cell(row=0, column=23, value="Chrom")
    tier4_fail_sheet.cell(row=0, column=24, value="Start")
    tier4_fail_sheet.cell(row=0, column=25, value="End")
    tier4_fail_sheet.cell(row=0, column=26, value="rsIDs")

    col = 27
    if 'mutect' in callers:
        tier4_fail_sheet.cell(row=0, column=col, value="MuTect_AF")
        col += 1

    if 'vardict' in callers:
        tier4_fail_sheet.cell(row=0, column=col, value="VarDict_AF")
        col += 1

    if 'freebayes' in callers:
        tier4_fail_sheet.cell(row=0, column=col, value="FreeBayes_AF")
        col += 1

    if 'scalpel' in callers:
        tier4_fail_sheet.cell(row=0, column=col, value="Scalpel_AF")
        col += 1

    if 'platypus' in callers:
        tier4_fail_sheet.cell(row=0, column=col, value="Platypus_AF")
        col += 1

    if 'pindel' in callers:
        tier4_fail_sheet.cell(row=0, column=col, value="Pindel_AF")
        col += 1

    row = 1
    for variant in report_data['']:
        tier4_fail_sheet.cell(row=row, column=0, value="{}".format(variant.sample))
        tier4_fail_sheet.cell(row=row, column=1, value="{}".format(variant.library_name))
        tier4_fail_sheet.cell(row=row, column=2, value="{}".format(variant.gene))
        tier4_fail_sheet.cell(row=row, column=3, value="{}".format(variant.amplicon_data['amplicon']))
        tier4_fail_sheet.cell(row=row, column=4, value="{}".format(variant.ref))
        tier4_fail_sheet.cell(row=row, column=5, value="{}".format(variant.alt))
        tier4_fail_sheet.cell(row=row, column=6, value="{}".format(variant.codon_change))
        tier4_fail_sheet.cell(row=row, column=7, value="{}".format(variant.aa_change))
        tier4_fail_sheet.cell(row=row, column=8, value="{}".format(variant.max_som_aaf))
        tier4_fail_sheet.cell(row=row, column=9, value="{}".format(",".join(variant.callers) or None))
        tier4_fail_sheet.cell(row=row, column=10, value="{}".format(",".join(variant.cosmic_ids) or None))
        tier4_fail_sheet.cell(row=row, column=11, value="{}".format(variant.cosmic_data['num_samples']))
        tier4_fail_sheet.cell(row=row, column=12, value="{}".format(variant.cosmic_data['aa']))
        tier4_fail_sheet.cell(row=row, column=13, value="{}".format(variant.clinvar_data['significance']))
        tier4_fail_sheet.cell(row=row, column=14, value="{}".format(variant.clinvar_data['hgvs']))
        tier4_fail_sheet.cell(row=row, column=15, value="{}".format(variant.clinvar_data['disease']))
        tier4_fail_sheet.cell(row=row, column=16,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['mean_coverage']))
        tier4_fail_sheet.cell(row=row, column=17,
                         value="{}".format(report_data['coverage'][variant.amplicon_data['amplicon']]['num_reads']))
        tier4_fail_sheet.cell(row=row, column=18, value="{}".format(variant.impact))
        tier4_fail_sheet.cell(row=row, column=19, value="{}".format(variant.severity))
        tier4_fail_sheet.cell(row=row, column=20, value="{}".format(variant.max_maf_all))
        tier4_fail_sheet.cell(row=row, column=21, value="{}".format(variant.min_depth))
        tier4_fail_sheet.cell(row=row, column=22, value="{}".format(variant.max_depth))
        tier4_fail_sheet.cell(row=row, column=23, value="{}".format(variant.chr))
        tier4_fail_sheet.cell(row=row, column=24, value="{}".format(variant.pos))
        tier4_fail_sheet.cell(row=row, column=25, value="{}".format(variant.end))
        tier4_fail_sheet.cell(row=row, column=26, value="{}".format(",".join(variant.rs_ids)))

        col = 27
        if 'mutect' in callers:
            tier4_fail_sheet.cell(row=row, column=col, value="{}".format(variant.mutect.get('AAF') or None))
            col += 1

        if 'vardict' in callers:
            tier4_fail_sheet.cell(row=row, column=col, value="{}".format(variant.vardict.get('AAF') or None))
            col += 1

        if 'freebayes' in callers:
            tier4_fail_sheet.cell(row=row, column=col, value="{}".format(variant.freebayes.get('AAF') or None))
            col += 1

        if 'scalpel' in callers:
            tier4_fail_sheet.cell(row=row, column=col, value="{}".format(variant.scalpel.get('AAF') or None))
            col += 1

        if 'platypus' in callers:
            tier4_fail_sheet.cell(row=row, column=col, value="{}".format(variant.platypus.get('AAF') or None))
            col += 1

        if 'pindel' in callers:
            tier4_fail_sheet.cell(row=row, column=col, value="{}".format(variant.pindel.get('AAF') or None))
            col += 1

        row += 1

    ####################################################################################################################

    wb.save(report_name)
